import os, json, sqlite3, uuid
from datetime import datetime, timedelta
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, g, send_file, flash)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import traceback  

from dotenv import load_dotenv
from s3_helper import upload_to_s3, get_s3_url, delete_survey_photos

load_dotenv()


from s3_helper import upload_to_s3, get_s3_url, delete_survey_photos

def attach_photo_urls(survey):
    """Adds _url keys for all photo columns using S3 presigned URLs."""
    photo_cols = [
        'photo_filename', 'interior_photo', 'shelf_photo', 'posm_photo',
        'cooler_photo_visi', 'cooler_photo_bottle', 'cooler_photo_freezer', 'space_photo'
    ]
    for col in photo_cols:
        survey[f"{col}_url"] = get_s3_url(survey.get(col))
    return survey

# ──────────────────────────────────────────────
# App Config
# ──────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'heritage-survey-secret-2024')
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10 MB
DB_PATH = os.path.join('instance', 'survey.db')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('instance', exist_ok=True)

ALLOWED_EXT = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# ──────────────────────────────────────────────
# Database helpers
# ──────────────────────────────────────────────
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

def query(sql, args=(), one=False):
    cur = get_db().execute(sql, args)
    rv = cur.fetchall()
    return (rv[0] if rv else None) if one else rv

def execute(sql, args=()):
    db = get_db()
    cur = db.execute(sql, args)
    db.commit()
    return cur


def generate_survey_id(se_id):
    """
    Generates a survey ID in format: {Region}{SalesOffice}{6-digit-counter}
    e.g. TSHSO1000001, TS2HSO15000042
    Counter is global per Sales Office — all SEs under same SO share it.
    Uses a DB transaction to ensure no duplicates even under concurrent access.
    """
    user = row_to_dict(query(
        "SELECT region, sales_office FROM users WHERE id=?", [se_id], one=True
    ))

    region       = (user.get('region') or 'TS').strip().upper().replace('-', '').replace(' ', '')
    sales_office = (user.get('sales_office') or 'HSO1').strip().upper().replace('-', '').replace(' ', '')

    db = get_db()

    # INSERT or IGNORE seeds the counter row if first survey for this SO
    db.execute(
        "INSERT OR IGNORE INTO survey_counters (sales_office, last_number) VALUES (?, 0)",
        [sales_office]
    )
    # Atomic increment — no race condition
    db.execute(
        "UPDATE survey_counters SET last_number = last_number + 1 WHERE sales_office = ?",
        [sales_office]
    )
    db.commit()

    row = db.execute(
        "SELECT last_number FROM survey_counters WHERE sales_office = ?",
        [sales_office]
    ).fetchone()

    number = row['last_number']
    return f"{region}{sales_office}{number:06d}"



def row_to_dict(row):
    if row is None: return None
    return dict(row)

# ──────────────────────────────────────────────
# DB Init
# ──────────────────────────────────────────────
def init_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        username    TEXT UNIQUE NOT NULL,
        password    TEXT NOT NULL,
        full_name   TEXT NOT NULL,
        employee_id TEXT,
        role        TEXT DEFAULT 'se',
        region      TEXT,
        sales_office TEXT,          -- e.g. HSO1, HSO5, HSO15
        is_active   INTEGER DEFAULT 1,
        created_at  TEXT DEFAULT (datetime('now')),
        last_login  TEXT
    );

    CREATE TABLE IF NOT EXISTS surveys (
        id              TEXT PRIMARY KEY,   -- e.g. TSHSO1000001
        se_id           INTEGER NOT NULL,
        status          TEXT DEFAULT 'draft',
        step_reached    INTEGER DEFAULT 1,
        outlet_name     TEXT,
        owner_name      TEXT,
        mobile          TEXT,
        whatsapp_enabled INTEGER DEFAULT 1,
        whatsapp_number TEXT,
        area            TEXT,
        latitude        REAL,
        longitude       REAL,
        photo_filename  TEXT,
        outlet_type     TEXT,
        industry_data   TEXT DEFAULT '{}',
        heritage_data   TEXT DEFAULT '{}',
        icecream_available TEXT,
        icecream_interest  TEXT,
        icecream_brands    TEXT,
        competition_data TEXT DEFAULT '{}',
        supply_data     TEXT DEFAULT '{}',
        cooling_data    TEXT DEFAULT '{}',
        se_remarks      TEXT,
        created_at      TEXT DEFAULT (datetime('now')),
        updated_at      TEXT DEFAULT (datetime('now')),
        submitted_at    TEXT,
        FOREIGN KEY (se_id) REFERENCES users(id)
    );

    -- One row per Sales Office, atomically incremented
    CREATE TABLE IF NOT EXISTS survey_counters (
        sales_office  TEXT PRIMARY KEY,
        last_number   INTEGER DEFAULT 0
    );
    """)

    # All existing ALTER TABLE blocks stay here unchanged
    for col in ['fssai_number']:
        try:
            db.execute(f"ALTER TABLE surveys ADD COLUMN {col} TEXT")
        except sqlite3.OperationalError:
            pass

    for col in ['interior_photo', 'cooler_photo', 'shelf_photo', 'posm_photo',
                'space_photo', 'cooler_photo_visi', 'cooler_photo_bottle', 'cooler_photo_freezer']:
        try:
            db.execute(f"ALTER TABLE surveys ADD COLUMN {col} TEXT")
        except sqlite3.OperationalError:
            pass

    # Add sales_office to users if upgrading existing DB
    try:
        db.execute("ALTER TABLE users ADD COLUMN sales_office TEXT")
    except sqlite3.OperationalError:
        pass

    db.commit()

    # Seed users
    existing = db.execute("SELECT id FROM users WHERE role='admin'").fetchone()
    if not existing:
        db.execute(
            "INSERT INTO users (username,password,full_name,employee_id,role) VALUES (?,?,?,?,?)",
            ('admin', generate_password_hash('admin123'), 'Admin User', 'ADM001', 'admin')
        )
        db.execute(
            "INSERT INTO users (username,password,full_name,employee_id,role,region,sales_office) VALUES (?,?,?,?,?,?,?)",
            ('se001', generate_password_hash('se001'), 'Ravi Kumar', 'SE001', 'se', 'TS', 'HSO1')
        )
        db.commit()
    db.close()
# ──────────────────────────────────────────────
# Auth helpers
# ──────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def current_user():
    if 'user_id' not in session:
        return None
    return row_to_dict(query("SELECT * FROM users WHERE id=?", [session['user_id']], one=True))

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT

# ──────────────────────────────────────────────
# Routes — Auth
# ──────────────────────────────────────────────
@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('se_dashboard'))

@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        user = row_to_dict(query("SELECT * FROM users WHERE username=? AND is_active=1",
                                 [username], one=True))
        if user and check_password_hash(user['password'], password):
            session.permanent = True
            app.permanent_session_lifetime = timedelta(hours=12)
            session['user_id'] = user['id']
            session['role']    = user['role']
            session['name']    = user['full_name']
            execute("UPDATE users SET last_login=datetime('now') WHERE id=?", [user['id']])
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('se_dashboard'))
        error = 'Invalid username or password.'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ──────────────────────────────────────────────
# Routes — SE Dashboard
# ──────────────────────────────────────────────
@app.route('/se')
@login_required
def se_dashboard():
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    user = current_user()
    surveys = query("""
        SELECT id, outlet_name, area, outlet_type, status, step_reached,
               created_at, updated_at, submitted_at
        FROM surveys WHERE se_id=? ORDER BY updated_at DESC
    """, [user['id']])
    surveys = [dict(s) for s in surveys]
    stats = {
        'total': len(surveys),
        'submitted': sum(1 for s in surveys if s['status']=='submitted'),
        'draft': sum(1 for s in surveys if s['status']=='draft'),
    }
    return render_template('se_dashboard.html', user=user, surveys=surveys, stats=stats)

# ──────────────────────────────────────────────
# Routes — Survey CRUD
# ──────────────────────────────────────────────
@app.route('/survey/new')
@login_required
def new_survey():
    survey_id = generate_survey_id(session['user_id'])
    execute("INSERT INTO surveys (id, se_id) VALUES (?, ?)",
            [survey_id, session['user_id']])
    return redirect(url_for('survey_step', survey_id=survey_id, step=1))

@app.route('/survey/<survey_id>/step/<int:step>', methods=['GET','POST'])
@login_required
def survey_step(survey_id, step):
    survey = row_to_dict(query("SELECT * FROM surveys WHERE id=?", [survey_id], one=True))
    if not survey:
        flash('Survey not found.', 'error')
        return redirect(url_for('se_dashboard'))
    if session.get('role') != 'admin' and survey['se_id'] != session['user_id']:
        return redirect(url_for('se_dashboard'))
    if survey['status'] == 'submitted' and session.get('role') != 'admin':
        flash('This survey has already been submitted. Contact admin to edit.', 'warning')
        return redirect(url_for('se_dashboard'))

    if request.method == 'POST':
        _save_step(survey_id, step, request)
        next_step = step + 1
        if next_step > 7:
            execute("""UPDATE surveys SET status='submitted', submitted_at=datetime('now'),
                       updated_at=datetime('now') WHERE id=?""", [survey_id])
            flash('Survey submitted successfully! 🎉', 'success')
            return redirect(url_for('se_dashboard'))
        execute("UPDATE surveys SET step_reached=MAX(step_reached,?), updated_at=datetime('now') WHERE id=?",
                [next_step, survey_id])
        return redirect(url_for('survey_step', survey_id=survey_id, step=next_step))

    survey = row_to_dict(query("SELECT * FROM surveys WHERE id=?", [survey_id], one=True))
    for json_col in ['industry_data','heritage_data','competition_data','supply_data','cooling_data']:
        try:
            survey[json_col] = json.loads(survey.get(json_col) or '{}')
        except: survey[json_col] = {}
    survey = attach_photo_urls(survey) 
    user = current_user()
    return render_template('survey.html', survey=survey, step=step, total_steps=7, user=user)

def _save_step(survey_id, step, req):
    f = req.form
    if step == 1:
        # Only process a new file upload if one was actually sent
        # (background upload already saved it to DB, so file field may be empty)
        photo_filename = None
        file = req.files.get('photo')
        if file and file.filename and allowed_file(file.filename):
            ext = file.filename.rsplit('.',1)[1].lower()
            photo_filename = f"outlet_{survey_id[:8]}_{uuid.uuid4().hex[:6]}.{ext}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], photo_filename))

        # If no new file in this request, keep whatever is in DB
        # (covers both: background-uploaded photo, and previously saved photo)
        if not photo_filename:
            existing = query("SELECT photo_filename FROM surveys WHERE id=?", [survey_id], one=True)
            photo_filename = existing['photo_filename'] if existing else None

        execute("""UPDATE surveys SET
            outlet_name=?, owner_name=?, mobile=?, whatsapp_enabled=?,
            whatsapp_number=?, area=?, latitude=?, longitude=?, photo_filename=?,
            fssai_number=?, updated_at=datetime('now') WHERE id=?""",
            [f.get('outlet_name'), f.get('owner_name'), f.get('mobile'),
             1 if f.get('whatsapp_enabled')=='1' else 0,
             f.get('whatsapp_number'), f.get('area'),
             f.get('latitude') or None, f.get('longitude') or None,
             photo_filename, f.get('fssai_number'), survey_id])
    elif step == 2:
        execute("UPDATE surveys SET outlet_type=?, updated_at=datetime('now') WHERE id=?",
                [f.get('outlet_type'), survey_id])
    elif step == 3:
        products = ['milk_10','milk_other','curd_10','curd_other','paneer','ghee','buttermilk','icecream']
        ind = {p: f.get(f'ind_{p}','') for p in products}
        her = {p: f.get(f'her_{p}','') for p in products}
        execute("""UPDATE surveys SET industry_data=?, heritage_data=?,
            icecream_available=?, icecream_interest=?, icecream_brands=?,
            updated_at=datetime('now') WHERE id=?""",
            [json.dumps(ind), json.dumps(her),
             f.get('icecream_available'), f.get('icecream_interest'),
             f.get('icecream_brands'), survey_id])
    elif step == 4:
        comp = {
            'milk_brands':   f.getlist('milk_brands'),
            'milk_rank':     f.getlist('milk_rank'),
            'curd_brands':   f.getlist('curd_brands'),
            'curd_rank':     f.getlist('curd_rank'),
            'paneer_brands': f.getlist('paneer_brands'),
            'paneer_rank':   f.getlist('paneer_rank'),
        }
        execute("UPDATE surveys SET competition_data=?, updated_at=datetime('now') WHERE id=?",
                [json.dumps(comp), survey_id])
    elif step == 5:
        sup = {
            'direct_supply':   f.get('direct_supply'),
            'customer_id':     f.get('customer_id'),
            'fresh_channels':  f.getlist('fresh_channels'),
            'fresh_reason':    f.get('fresh_reason'),
            'other_channels':  f.getlist('other_channels'),
            'other_reason':    f.get('other_reason'),
            'ic_distributor':  f.get('ic_distributor'),
        }
        execute("UPDATE surveys SET supply_data=?, updated_at=datetime('now') WHERE id=?",
                [json.dumps(sup), survey_id])
    elif step == 6:
        cool = {
            'assets':           f.getlist('cooling_assets'),
            'ownership':        f.get('cooler_ownership'),
            'interested':       f.get('cooler_interest'),
            'preferred_type':   f.get('cooler_type'),
            'space_available':  f.get('space_available'),
        }
        execute("UPDATE surveys SET cooling_data=?, updated_at=datetime('now') WHERE id=?",
                [json.dumps(cool), survey_id])
    elif step == 7:
        execute("UPDATE surveys SET se_remarks=?, updated_at=datetime('now') WHERE id=?",
                [f.get('se_remarks'), survey_id])

# ──────────────────────────────────────────────
# API — Background photo upload (NEW)
# Compresses on client, uploads here immediately on photo select
# so "Next" button doesn't have to wait for the upload.
# ──────────────────────────────────────────────
@app.route('/api/survey/<survey_id>/upload-photo', methods=['POST'])
@login_required
def upload_photo_bg(survey_id):
    if session.get('role') != 'admin':
        survey = query("SELECT se_id FROM surveys WHERE id=?", [survey_id], one=True)
        if not survey or survey['se_id'] != session['user_id']:
            return jsonify({'error': 'Unauthorized'}), 403

    file = request.files.get('photo')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400

    allowed_types = {'photo', 'interior_photo', 'cooler_photo', 'shelf_photo',
                     'posm_photo', 'space_photo', 'cooler_photo_visi',
                     'cooler_photo_bottle', 'cooler_photo_freezer'}
    photo_type = request.form.get('photo_type', 'photo')
    if photo_type not in allowed_types:
        photo_type = 'photo'

    col_name = 'photo_filename' if photo_type == 'photo' else photo_type

    try:
        s3_key = upload_to_s3(file.stream, survey_id, photo_type)
        execute(f"UPDATE surveys SET {col_name}=?, updated_at=datetime('now') WHERE id=?",
                [s3_key, survey_id])
        return jsonify({'ok': True, 'filename': s3_key})
    except Exception as e:
        traceback.print_exc() 
        return jsonify({'error': str(e)}), 500
    

@app.route('/survey/<survey_id>/delete', methods=['POST'])
@login_required
def delete_survey(survey_id):
    survey = row_to_dict(query("SELECT * FROM surveys WHERE id=?", [survey_id], one=True))
    if survey and (survey['se_id'] == session['user_id'] or session.get('role') == 'admin'):
        delete_survey_photos(survey_id)   # ← delete from S3
        execute("DELETE FROM surveys WHERE id=?", [survey_id])
        flash('Survey deleted.', 'info')
    return redirect(url_for('se_dashboard') if session.get('role') != 'admin' else url_for('admin_surveys'))


@app.route('/survey/<survey_id>/reopen', methods=['POST'])
@admin_required
def reopen_survey(survey_id):
    execute("UPDATE surveys SET status='draft', submitted_at=NULL WHERE id=?", [survey_id])
    flash('Survey reopened for editing.', 'success')
    return redirect(url_for('admin_surveys'))

# ──────────────────────────────────────────────
# Routes — Admin
# ──────────────────────────────────────────────
@app.route('/admin')
@admin_required
def admin_dashboard():
    total_surveys   = query("SELECT COUNT(*) c FROM surveys", one=True)['c']
    submitted       = query("SELECT COUNT(*) c FROM surveys WHERE status='submitted'", one=True)['c']
    draft           = query("SELECT COUNT(*) c FROM surveys WHERE status='draft'", one=True)['c']
    total_se        = query("SELECT COUNT(*) c FROM users WHERE role='se' AND is_active=1", one=True)['c']
    active_today    = query("SELECT COUNT(DISTINCT se_id) c FROM surveys WHERE date(updated_at)=date('now')", one=True)['c']

    se_activity = query("""
        SELECT u.full_name, u.employee_id, u.region, u.last_login,
               COUNT(s.id) total, SUM(s.status='submitted') submitted,
               SUM(s.status='draft') draft,
               MAX(s.updated_at) last_activity
        FROM users u
        LEFT JOIN surveys s ON u.id=s.se_id
        WHERE u.role='se' AND u.is_active=1
        GROUP BY u.id ORDER BY last_activity DESC
    """)
    se_activity = [dict(r) for r in se_activity]

    recent = query("""
        SELECT s.id, s.outlet_name, s.area, s.outlet_type, s.status,
               s.step_reached, s.updated_at, u.full_name se_name
        FROM surveys s JOIN users u ON s.se_id=u.id
        ORDER BY s.updated_at DESC LIMIT 10
    """)
    recent = [dict(r) for r in recent]

    type_breakdown = query("""
        SELECT outlet_type, COUNT(*) cnt FROM surveys
        WHERE outlet_type IS NOT NULL GROUP BY outlet_type ORDER BY cnt DESC LIMIT 10
    """)
    type_breakdown = [dict(r) for r in type_breakdown]

    daily_trend = query("""
        SELECT date(submitted_at) day, COUNT(*) cnt
        FROM surveys WHERE submitted_at > date('now','-7 days')
        GROUP BY day ORDER BY day
    """)
    daily_trend = [dict(r) for r in daily_trend]

    return render_template('admin_dashboard.html',
        stats={'total': total_surveys, 'submitted': submitted, 'draft': draft,
               'total_se': total_se, 'active_today': active_today,
               'completion_rate': round(submitted/total_surveys*100) if total_surveys else 0},
        se_activity=se_activity, recent=recent,
        type_breakdown=type_breakdown, daily_trend=daily_trend,
        user=current_user())


@app.route('/admin/users/<int:user_id>/edit', methods=['POST'])
@admin_required
def edit_user(user_id):
    full_name    = request.form.get('full_name','').strip()
    employee_id  = request.form.get('employee_id','').strip()
    region       = request.form.get('region','').strip()
    sales_office = request.form.get('sales_office','').strip()   # ← ADD
    execute(
        "UPDATE users SET full_name=?, employee_id=?, region=?, sales_office=? WHERE id=?",
        [full_name, employee_id, region, sales_office, user_id]
    )
    flash('SE details updated successfully.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    user = query("SELECT role FROM users WHERE id=?", [user_id], one=True)
    if user and user['role'] != 'admin':
        execute("DELETE FROM users WHERE id=?", [user_id])
        flash('SE deleted completely.', 'info')
    return redirect(url_for('admin_users'))


# AFTER
@app.route('/admin/surveys')
@admin_required
def admin_surveys():
    se_filter      = request.args.get('se', '')
    status_filter  = request.args.get('status', '')
    date_from      = request.args.get('from', '')
    date_to        = request.args.get('to', '')
    region_filter  = request.args.get('region', '')
    so_filter      = request.args.get('sales_office', '')

    sql = """SELECT s.*, u.full_name se_name, u.employee_id se_empid, u.region, u.sales_office
             FROM surveys s JOIN users u ON s.se_id=u.id WHERE 1=1"""
    args = []
    if se_filter:      sql += " AND u.id=?";                args.append(se_filter)
    if status_filter:  sql += " AND s.status=?";            args.append(status_filter)
    if date_from:      sql += " AND date(s.created_at)>=?"; args.append(date_from)
    if date_to:        sql += " AND date(s.created_at)<=?"; args.append(date_to)
    if region_filter:  sql += " AND u.region=?";            args.append(region_filter)
    if so_filter:      sql += " AND u.sales_office=?";      args.append(so_filter)
    sql += " ORDER BY s.updated_at DESC"

    surveys = [dict(r) for r in query(sql, args)]

    all_ses = [dict(r) for r in query("""
        SELECT id, full_name, employee_id, region, sales_office
        FROM users WHERE role='se' AND is_active=1 ORDER BY full_name
    """)]
    regions = sorted(set(s['region'] for s in all_ses if s['region']))
    offices = sorted(set(s['sales_office'] for s in all_ses if s['sales_office']))

    return render_template('admin_surveys.html',
        surveys=surveys,
        se_list=all_ses,
        regions=regions,
        offices=offices,
        all_ses_json=json.dumps(all_ses),
        filters={'se': se_filter, 'status': status_filter,
                 'from': date_from, 'to': date_to,
                 'region': region_filter, 'sales_office': so_filter},
        user=current_user())

@app.route('/admin/survey/<survey_id>/view')
@admin_required
def admin_view_survey(survey_id):
    survey = row_to_dict(query("""
        SELECT s.*, u.full_name se_name, u.employee_id, u.region
        FROM surveys s JOIN users u ON s.se_id=u.id WHERE s.id=?
    """, [survey_id], one=True))
    if not survey: return redirect(url_for('admin_surveys'))
    for col in ['industry_data','heritage_data','competition_data','supply_data','cooling_data']:
        try:    survey[col] = json.loads(survey.get(col) or '{}')
        except: survey[col] = {}
    survey = attach_photo_urls(survey)   # ← add this
    return render_template('admin_view_survey.html', survey=survey, user=current_user())


@app.route('/admin/users')
@admin_required
def admin_users():
    users = [dict(r) for r in query("""
        SELECT u.*, COUNT(s.id) total_surveys, SUM(s.status='submitted') submitted
        FROM users u LEFT JOIN surveys s ON u.id=s.se_id
        WHERE u.role='se' GROUP BY u.id ORDER BY u.created_at DESC
    """)]
    return render_template('admin_users.html', users=users, user=current_user())

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def add_user():
    username     = request.form.get('username','').strip()
    password     = request.form.get('password','')
    full_name    = request.form.get('full_name','').strip()
    employee_id  = request.form.get('employee_id','').strip()
    region       = request.form.get('region','').strip()
    sales_office = request.form.get('sales_office','').strip()   # ← ADD

    if not username or not password or not full_name:
        flash('Username, password and name are required.', 'error')
        return redirect(url_for('admin_users'))

    existing = query("SELECT id FROM users WHERE username=?", [username], one=True)
    if existing:
        flash(f'Username "{username}" already exists.', 'error')
        return redirect(url_for('admin_users'))

    execute(
        "INSERT INTO users (username,password,full_name,employee_id,role,region,sales_office) VALUES (?,?,?,?,?,?,?)",
        [username, generate_password_hash(password), full_name, employee_id, 'se', region, sales_office]
    )
    flash(f'SE "{full_name}" added successfully.', 'success')
    return redirect(url_for('admin_users'))



@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def toggle_user(user_id):
    user = row_to_dict(query("SELECT * FROM users WHERE id=?", [user_id], one=True))
    if user and user['role'] != 'admin':
        new_status = 0 if user['is_active'] else 1
        execute("UPDATE users SET is_active=? WHERE id=?", [new_status, user_id])
        status_text = 'activated' if new_status else 'deactivated'
        flash(f'User {status_text}.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def reset_password(user_id):
    new_pw = request.form.get('new_password','')
    if len(new_pw) < 4:
        flash('Password must be at least 4 characters.', 'error')
    else:
        execute("UPDATE users SET password=? WHERE id=?",
                [generate_password_hash(new_pw), user_id])
        flash('Password reset successfully.', 'success')
    return redirect(url_for('admin_users'))

# ──────────────────────────────────────────────
# Excel Export
# ──────────────────────────────────────────────
@app.route('/admin/export')
@admin_required
def export_excel():
    surveys = [dict(r) for r in query("""
        SELECT s.*, u.full_name se_name, u.employee_id se_empid, u.region
        FROM surveys s JOIN users u ON s.se_id=u.id
        ORDER BY s.submitted_at DESC
    """)]

    def photo_url(s3_key):
        if not s3_key:
            return ''
        return get_s3_url(s3_key) or '' 

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Survey Summary"

    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="F1F8E9")
    green_fill  = PatternFill("solid", fgColor="C8E6C9")
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    headers = [
        # Identity
        "Survey ID", "SE Name", "Employee ID", "Region", "Status",
        # Step 1 — Outlet basics
        "Outlet Name", "Owner Name", "Mobile", "WhatsApp Enabled",
        "WhatsApp Number", "FSSAI Number", "Area",
        "Latitude", "Longitude",
        # Step 1 — Photo
        "Outlet Photo",
        # Step 2 — Classification + Photos
        "Outlet Type",
        "Interior Photo", "Heritage Shelf Photo", "POSM Photo",
        # Step 3 — Industry
        "Milk ₹10 (LPD)", "Milk Other (LPD)",
        "Curd ₹10 (Kg/D)", "Curd Other (Kg/D)",
        "Paneer (Kg/Wk)", "Ghee (Kg/Mo)",
        "Buttermilk (Units/D)", "Ice Cream (₹/Mo)",
        "IC Available", "IC Interest", "IC Brands",
        # Step 3 — Heritage
        "Heritage Milk ₹10", "Heritage Milk Other",
        "Heritage Curd ₹10", "Heritage Curd Other",
        "Heritage Paneer", "Heritage Ghee",
        "Heritage Buttermilk", "Heritage Ice Cream",
        # Step 4 — Competition
        "Milk Brands (Top3)", "Curd Brands (Top3)", "Paneer Brands (Top3)",
        # Step 5 — Supply
        "Direct Supply", "Customer ID",
        "Fresh Channels", "Fresh No-Supply Reason",
        "Other Channels", "Other No-Supply Reason",
        "IC Distributor",
        # Step 6 — Cooling
        "Cooling Assets", "Cooler Ownership",
        "Space Available", "Cooler Interest", "Preferred Cooler Type",
        # Step 6 — Cooler Photos
        "Visi Cooler Photo", "Bottle Cooler Photo", "IC Freezer Photo",
        "Space Photo",
        # Step 7
        "SE Remarks", "Submitted At", "Created At"
    ]

    ws1.row_dimensions[1].height = 22
    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
        ws1.column_dimensions[get_column_letter(ci)].width = max(12, min(30, len(h) + 4))

    # BEFORE
    BASE_URL = "http://localhost:5000/static/uploads/"

    def photo_url(filename):
        return (BASE_URL + filename) if filename else ''

    # AFTER
    def photo_url(s3_key):
        if not s3_key:
            return ''
        if '/' not in s3_key:   # old local filename — skip
            return ''
        return get_s3_url(s3_key) or ''

    for ri, sv in enumerate(surveys, 2):
        ind  = json.loads(sv.get('industry_data')   or '{}')
        her  = json.loads(sv.get('heritage_data')   or '{}')
        comp = json.loads(sv.get('competition_data') or '{}')
        sup  = json.loads(sv.get('supply_data')     or '{}')
        cool = json.loads(sv.get('cooling_data')    or '{}')

        row = [
            # Identity
            sv.get('id',''),
            sv.get('se_name',''),
            sv.get('se_empid',''),
            sv.get('region',''),
            sv.get('status','').upper(),
            # Step 1
            sv.get('outlet_name',''),
            sv.get('owner_name',''),
            sv.get('mobile',''),
            'Yes' if sv.get('whatsapp_enabled') else 'No',
            sv.get('whatsapp_number',''),
            sv.get('fssai_number',''),
            sv.get('area',''),
            sv.get('latitude',''),
            sv.get('longitude',''),
            photo_url(sv.get('photo_filename','')),
            # Step 2
            sv.get('outlet_type',''),
            photo_url(sv.get('interior_photo','')),
            photo_url(sv.get('shelf_photo','')),
            photo_url(sv.get('posm_photo','')),
            # Step 3 — Industry
            ind.get('milk_10',''),
            ind.get('milk_other',''),
            ind.get('curd_10',''),
            ind.get('curd_other',''),
            ind.get('paneer',''),
            ind.get('ghee',''),
            ind.get('buttermilk',''),
            ind.get('icecream',''),
            sv.get('icecream_available',''),
            sv.get('icecream_interest',''),
            sv.get('icecream_brands',''),
            # Step 3 — Heritage
            her.get('milk_10',''),
            her.get('milk_other',''),
            her.get('curd_10',''),
            her.get('curd_other',''),
            her.get('paneer',''),
            her.get('ghee',''),
            her.get('buttermilk',''),
            her.get('icecream',''),
            # Step 4
            ', '.join(comp.get('milk_rank',[])),
            ', '.join(comp.get('curd_rank',[])),
            ', '.join(comp.get('paneer_rank',[])),
            # Step 5
            sup.get('direct_supply',''),
            sup.get('customer_id',''),
            ', '.join(sup.get('fresh_channels',[])),
            sup.get('fresh_reason',''),
            ', '.join(sup.get('other_channels',[])),
            sup.get('other_reason',''),
            sup.get('ic_distributor',''),
            # Step 6
            ', '.join(cool.get('assets',[])),
            cool.get('ownership',''),
            cool.get('space_available',''),
            cool.get('interested',''),
            cool.get('preferred_type',''),
            photo_url(sv.get('cooler_photo_visi','')),
            photo_url(sv.get('cooler_photo_bottle','')),
            photo_url(sv.get('cooler_photo_freezer','')),
            photo_url(sv.get('space_photo','')),
            # Step 7
            sv.get('se_remarks',''),
            sv.get('submitted_at',''),
            sv.get('created_at',''),
        ]

        fill = green_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(ri, ci, val)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(vertical='center')
            # Status column bold + coloured
            if ci == 5:
                cell.font = Font(bold=True,
                    color="1B5E20" if str(val) == 'SUBMITTED' else "E65100")
            # Photo URL columns — make them clickable hyperlinks
            photo_cols = {15, 17, 18, 19, 52, 53, 54, 55}  # col indices of photo_url fields
            if ci in photo_cols and val:
                cell.hyperlink = val
                cell.font = Font(color="1155CC", underline="single")
                cell.value = "View Photo"

    ws1.freeze_panes = 'A2'

    # ── Sheet 2: SE Performance (unchanged) ──
    ws2 = wb.create_sheet("SE Performance")
    se_perf = query("""
        SELECT u.full_name, u.employee_id, u.region, u.last_login,
               COUNT(s.id) total, SUM(s.status='submitted') submitted,
               SUM(s.status='draft') drafts
        FROM users u LEFT JOIN surveys s ON u.id=s.se_id
        WHERE u.role='se' GROUP BY u.id
    """)
    h2 = ["SE Name","Employee ID","Region","Last Login",
          "Total Surveys","Submitted","Drafts","Completion %"]
    ws2.row_dimensions[1].height = 22
    for ci, h in enumerate(h2, 1):
        cell = ws2.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        ws2.column_dimensions[get_column_letter(ci)].width = 18
    for ri, row in enumerate(se_perf, 2):
        r = dict(row)
        total = r['total'] or 0
        sub   = r['submitted'] or 0
        pct   = round(sub / total * 100) if total else 0
        vals  = [r['full_name'], r['employee_id'], r['region'], r['last_login'],
                 total, sub, r['drafts'] or 0, f"{pct}%"]
        fill  = alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(ri, ci, v)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(vertical='center')
    ws2.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Heritage_Outlet_Survey_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=fname)



# ──────────────────────────────────────────────
# API — Autosave
# ──────────────────────────────────────────────
@app.route('/api/survey/<survey_id>/autosave', methods=['POST'])
@login_required
def autosave(survey_id):
    data = request.get_json()
    step = data.get('step', 1)
    if session.get('role') != 'admin':
        survey = query("SELECT se_id FROM surveys WHERE id=?", [survey_id], one=True)
        if not survey or survey['se_id'] != session['user_id']:
            return jsonify({'error': 'Unauthorized'}), 403
    execute("UPDATE surveys SET updated_at=datetime('now'), step_reached=MAX(step_reached,?) WHERE id=?",
            [step, survey_id])
    return jsonify({'ok': True})

# ──────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    print("\n✅ Heritage Survey App Starting...")
    print("─────────────────────────────────")
    print("  Admin login : admin / admin123")
    print("  SE login    : se001 / se001")
    print("  URL         : http://localhost:5000")
    print("─────────────────────────────────\n")
    app.run(debug=True, host='0.0.0.0', port=5000)